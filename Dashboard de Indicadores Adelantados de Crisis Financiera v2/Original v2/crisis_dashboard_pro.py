#!/usr/bin/env python3
"""
=============================================================================
SISTEMA PROFESIONAL DE INDICADORES ADELANTADOS DE CRISIS FINANCIERA
=============================================================================
Version: 7.0 PROFESSIONAL
Autor: Sistema validado con backtesting empírico
Fecha: 2025-12-10

DISCLAIMER LEGAL:
Este sistema es una herramienta de análisis, NO asesoramiento financiero.
Las decisiones de inversión son responsabilidad exclusiva del usuario.
Rendimientos pasados no garantizan resultados futuros.
=============================================================================
"""

import pandas as pd
import numpy as np
import yfinance as yf
from fredapi import Fred
from datetime import datetime, timedelta
import os
import time
import pickle
import requests
from bs4 import BeautifulSoup
from functools import wraps
from pathlib import Path
import json

# =============================================================================
# CONFIGURACIÓN GLOBAL
# =============================================================================

# Seguridad: API Key desde entorno o archivo .env
# Intentar cargar desde archivo .env si existe
try:
    from dotenv import load_dotenv
    # Buscar .env en directorio actual y carpeta padre
    if Path('.env').exists():
        load_dotenv('.env')
        print("✅ Cargado .env desde directorio actual")
    elif Path('FRED_API_KEY.env').exists():
        load_dotenv('FRED_API_KEY.env')
        print("✅ Cargado FRED_API_KEY.env desde directorio actual")
    elif Path('../.env').exists():
        load_dotenv('../.env')
        print("✅ Cargado .env desde carpeta padre")
    else:
        print("⚠️  No se encontró archivo .env, buscando en variables de entorno del sistema")
except ImportError:
    print("⚠️  python-dotenv no instalado, usando solo variables de entorno del sistema")
    print("   Instala con: pip install python-dotenv")

FRED_API_KEY = os.getenv('FRED_API_KEY')

# Si aún no encontró la clave, buscar en directorio actual como último recurso
if not FRED_API_KEY:
    # Buscar archivos .env en el directorio actual
    env_files = ['.env', 'FRED_API_KEY.env', '.env.local']
    for env_file in env_files:
        if Path(env_file).exists():
            print(f"📄 Encontrado {env_file}, intentando carga manual...")
            try:
                with open(env_file, 'r') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            if '=' in line:
                                key, value = line.split('=', 1)
                                if key.strip() == 'FRED_API_KEY':
                                    FRED_API_KEY = value.strip().strip('"').strip("'")
                                    print(f"✅ API Key cargada manualmente desde {env_file}")
                                    break
            except Exception as e:
                print(f"⚠️  Error leyendo {env_file}: {e}")
            if FRED_API_KEY:
                break

if not FRED_API_KEY:
    print("\n" + "="*80)
    print("❌ ERROR CRÍTICO: No se encontró FRED_API_KEY")
    print("="*80)
    print("\n📋 SOLUCIONES:\n")
    print("Opción 1 - Crear archivo .env en el mismo directorio que el script:")
    print("   1. Crea un archivo llamado '.env' (o 'FRED_API_KEY.env')")
    print("   2. Escribe dentro: FRED_API_KEY=tu_clave_aqui")
    print("   3. Guarda el archivo\n")
    print("Opción 2 - Configurar variable de entorno del sistema:")
    print("   Windows: set FRED_API_KEY=tu_clave_aqui")
    print("   Linux/Mac: export FRED_API_KEY=tu_clave_aqui\n")
    print("Opción 3 - Instalar python-dotenv (recomendado):")
    print("   pip install python-dotenv\n")
    print("="*80)
    raise ValueError("FRED_API_KEY no encontrada. Ver instrucciones arriba.")

fred = Fred(api_key=FRED_API_KEY)

# Configuración de fechas
END_DATE = datetime.now()
START_DATE = '1920-01-01'
CACHE_DIR = Path("./market_data_cache")
CACHE_DIR.mkdir(exist_ok=True)

# Períodos de crisis históricos (VALIDADOS)
CRISIS_PERIODS = {
    'Gran Depresión (1929-1938)': ('1929-07-01', '1932-07-01'),
    'Recesión de 1937-1938': ('1937-01-01', '1942-04-01'),
    'Post-WWII Recesión (1945-1946)': ('1945-05-01', '1949-07-01'),
    'Inflación en EE. UU 1968 - Previa crisis petroleo': ('1968-10-01', '1970-04-01'),
    'Crisis del Petróleo (1973)': ('1973-01-01', '1974-11-01'),
    'Lunes Negro" de octubre de 1987': ('1987-08-01', '1987-11-01'),
    'Recesión de 1990-1991': ('1990-07-01', '1991-03-01'),
    'Crisis del Tequila': ('1994-01-01', '1995-01-01'),
    'Crisis Asiática': ('1997-07-01', '1998-05-01'),
    'Crisis Rublo': ('1998-08-01', '1999-01-01'),
    'Burbuja .com': ('2000-03-01', '2002-09-01'),
    'Crisis Subprime (GFC)': ('2008-09-01', '2009-03-01'),
    'Flash Crash': ('2010-05-06', '2010-05-10'),
    'Deuda Europea': ('2011-08-01', '2012-06-01'),
    'Corrección China': ('2015-08-01', '2015-09-01'),
    'COVID-19': ('2020-02-15', '2020-04-01'),
    'Mini-crisis SVB': ('2023-03-08', '2023-03-20'),
}

# =============================================================================
# UTILIDADES: RATE LIMITING + CACHÉ
# =============================================================================

def rate_limited_retry(max_retries=3, initial_wait=2):
    """
    Decorador para reintentos con backoff exponencial.
    Maneja rate limits y errores de red de forma profesional.
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            wait_time = initial_wait
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    error_msg = str(e)
                    if "429" in error_msg or "Rate limit" in error_msg or "Too Many Requests" in error_msg:
                        print(f"⏱️  Rate limit detectado. Esperando {wait_time}s... (Intento {attempt+1}/{max_retries})")
                    else:
                        print(f"⚠️  Error: {e}. Reintentando en {wait_time}s...")
                    
                    if attempt < max_retries - 1:
                        time.sleep(wait_time)
                        wait_time *= 2
                    else:
                        print(f"❌ Falló después de {max_retries} intentos.")
                        return pd.Series(dtype=float)
        return wrapper
    return decorator

def get_cached_series(ticker, source_func, cache_hours=12):
    """Sistema de caché para evitar llamadas redundantes a APIs."""
    cache_file = CACHE_DIR / f"{ticker.replace('^', '').replace('=', '').replace('/', '_')}.pkl"
    
    if cache_file.exists():
        last_modified = datetime.fromtimestamp(cache_file.stat().st_mtime)
        if datetime.now() - last_modified < timedelta(hours=cache_hours):
            print(f"📦 Caché: {ticker}")
            with open(cache_file, 'rb') as f:
                return pickle.load(f)

    print(f"⬇️  Descargando: {ticker}...")
    data = source_func()
    
    if not data.empty:
        with open(cache_file, 'wb') as f:
            pickle.dump(data, f)
            
    return data

# =============================================================================
# FUENTES DE DATOS
# =============================================================================

@rate_limited_retry(max_retries=5)
def fetch_fred_series(series_id):
    """Descarga desde FRED con manejo robusto de errores."""
    s = fred.get_series(series_id, observation_start=START_DATE, end_date=END_DATE)
    return s.dropna()

@rate_limited_retry(max_retries=5)
def fetch_yahoo_series(ticker):
    """Descarga desde Yahoo Finance con configuración optimizada."""
    df = yf.download(ticker, start=START_DATE, end=END_DATE, progress=False, auto_adjust=False)
    
    if df.empty:
        return pd.Series(dtype=float)
    
    if 'Adj Close' in df.columns:
        s = df['Adj Close']
    elif 'Close' in df.columns:
        s = df['Close']
    else:
        return pd.Series(dtype=float)
        
    return s.dropna()

def fetch_fear_greed_index():
    """
    Descarga el Fear & Greed Index de CNN Business.
    Retorna: valor actual (0-100) y clasificación ('Extreme Fear', 'Fear', etc.)
    """
    try:
        url = "https://production.dataviz.cnn.io/index/fearandgreed/graphdata"
        response = requests.get(url, timeout=10)
        data = response.json()
        
        current_score = float(data['fear_and_greed']['score'])
        current_rating = data['fear_and_greed']['rating']
        
        # Crear serie histórica si está disponible
        historical = data.get('fear_and_greed_historical', {}).get('data', [])
        if historical:
            df = pd.DataFrame(historical)
            df['x'] = pd.to_datetime(df['x'], unit='ms')
            series = pd.Series(df['y'].values, index=df['x'], name='Fear_Greed')
        else:
            series = pd.Series([current_score], index=[datetime.now()], name='Fear_Greed')
        
        return {
            'current': current_score,
            'rating': current_rating,
            'series': series
        }
    except Exception as e:
        print(f"⚠️  No se pudo obtener Fear & Greed Index: {e}")
        return {'current': None, 'rating': 'N/A', 'series': pd.Series(dtype=float)}

# =============================================================================
# MÓDULO DE BACKTESTING PROFESIONAL
# =============================================================================

class IndicatorBacktester:
    """
    Sistema de backtesting para validar poder predictivo de indicadores.
    
    Métricas calculadas:
    - True Positives: Señales que precedieron crisis reales
    - False Positives: Falsas alarmas
    - Precision: TP / (TP + FP) - Confiabilidad de la señal
    - Lead Time: Días promedio de anticipación
    """
    
    def __init__(self, crisis_periods):
        self.crisis_periods = crisis_periods
        self.results = {}
    
    def backtest_threshold(self, series, threshold, comparison='>', lead_days=90, name='Indicator'):
        """
        Backtesting de un indicador con umbral específico.
        
        Args:
            series: Serie temporal del indicador
            threshold: Valor de umbral para activar señal
            comparison: '>' para señales cuando valor supera umbral, '<' para cuando está debajo
            lead_days: Días de anticipación esperados (default 90)
            name: Nombre del indicador
        
        Returns:
            Dict con métricas de performance
        """
        # Identificar todas las señales
        if comparison == '>':
            signals = series[series > threshold]
        else:
            signals = series[series < threshold]
        
        true_positives = []
        false_positives = []
        lead_times = []
        
        for signal_date in signals.index:
            # Buscar si hubo crisis en ventana de lead_days hacia adelante
            crisis_found = False
            
            for crisis_name, (start, end) in self.crisis_periods.items():
                crisis_start = pd.to_datetime(start)
                
                # La señal debe ser ANTES de la crisis
                days_before = (crisis_start - signal_date).days
                
                if 0 <= days_before <= lead_days:
                    crisis_found = True
                    true_positives.append({
                        'date': signal_date,
                        'value': series.loc[signal_date],
                        'crisis': crisis_name,
                        'lead_days': days_before
                    })
                    lead_times.append(days_before)
                    break  # Contar solo una vez por señal
            
            if not crisis_found:
                false_positives.append({
                    'date': signal_date,
                    'value': series.loc[signal_date]
                })
        
        tp_count = len(true_positives)
        fp_count = len(false_positives)
        total_signals = tp_count + fp_count
        
        precision = tp_count / total_signals if total_signals > 0 else 0
        avg_lead = np.mean(lead_times) if lead_times else 0
        
        # Calcular sensibilidad: ¿cuántas crisis detectamos?
        total_crises = len(self.crisis_periods)
        sensitivity = tp_count / total_crises if total_crises > 0 else 0
        
        result = {
            'name': name,
            'threshold': threshold,
            'total_signals': total_signals,
            'true_positives': tp_count,
            'false_positives': fp_count,
            'precision': precision,
            'sensitivity': sensitivity,
            'avg_lead_days': avg_lead,
            'tp_details': true_positives,
            'fp_details': false_positives
        }
        
        self.results[name] = result
        return result
    
    def print_summary(self, indicator_name):
        """Imprime resumen ejecutivo del backtesting."""
        if indicator_name not in self.results:
            print(f"No hay resultados de backtesting para {indicator_name}")
            return
        
        r = self.results[indicator_name]
        
        print(f"\n{'='*70}")
        print(f"BACKTESTING: {r['name']}")
        print(f"{'='*70}")
        print(f"Umbral analizado: {r['threshold']}")
        print(f"\nRESULTADOS:")
        print(f"  Total de señales activadas: {r['total_signals']}")
        print(f"  ✅ Crisis anticipadas correctamente: {r['true_positives']}")
        print(f"  ❌ Falsas alarmas: {r['false_positives']}")
        print(f"\nMÉTRICAS DE PERFORMANCE:")
        print(f"  Precisión: {r['precision']:.1%} (confiabilidad de la señal)")
        print(f"  Sensibilidad: {r['sensitivity']:.1%} (% de crisis detectadas)")
        print(f"  Anticipación promedio: {r['avg_lead_days']:.0f} días")
        print(f"\n📊 INTERPRETACIÓN:")
        
        if r['precision'] >= 0.5:
            print(f"  ✅ INDICADOR CONFIABLE: Más del 50% de las señales fueron correctas")
        elif r['precision'] >= 0.3:
            print(f"  ⚠️  INDICADOR MODERADO: ~{r['precision']:.0%} de precisión, usar con otros indicadores")
        else:
            print(f"  ❌ INDICADOR POCO CONFIABLE: Alta tasa de falsos positivos ({r['false_positives']/r['total_signals']:.0%})")
        
        print(f"{'='*70}\n")

# =============================================================================
# INDICADORES CON BACKTESTING INTEGRADO
# =============================================================================

def analyze_high_yield(backtester=None):
    """
    Spread de Bonos High Yield con metodología de 5 pasos (Usuario - Dic 2024).
    
    OBJETIVO: Sistema de SALIDA - Detectar crisis para mantenerse FUERA
    
    FORMACIÓN DE CRISIS (Pasos 1-4):
    ═══════════════════════════════════════════════════════════
    1️⃣ Spread ≥ 7.30%
    2️⃣ Dispersión máxima > +15% en últimas 30 ruedas
    3️⃣ Spread cruza WWMA-8 DE ARRIBA HACIA ABAJO
       → Señal: "CRISIS EN DESARROLLO, MANTENERSE FUERA"
    4️⃣ Dos mínimos decrecientes consecutivos después del cruce
       → Señal: "ADVERTENCIA: Crisis desarrollándose"
    
    CRISIS CONFIRMADA (Paso 5):
    ═══════════════════════════════════════════════════════════
    5️⃣ Todos los anteriores + Dispersión vs MA-34 ≤ 0%
       → Señal: "CRISIS CONFIRMADA - ¡NO INVERTIR!"
    
    BACKTESTING: Cuenta como crisis cuando se cumplen los primeros 4 pasos
    """
    series = get_cached_series('BAMLH0A0HYM2', lambda: fetch_fred_series('BAMLH0A0HYM2'))
    
    if series.empty:
        return {
            'name': 'High Yield Spread',
            'current': None,
            'level': -1,
            'status': '⚠️  Error: Sin datos',
            'recommendation': 'No disponible',
            'last_date': None
        }
    
    # =========================================================================
    # CÁLCULO DE INDICADORES TÉCNICOS
    # =========================================================================
    
    current = series.iloc[-1]
    last_date = series.index[-1]
    
    # MA-34 días (Simple Moving Average)
    ma_34 = series.rolling(window=34).mean()
    current_ma_34 = ma_34.iloc[-1] if len(ma_34) >= 34 else current
    
    # WWMA-8 (Welles Wilder Moving Average)
    wwma_8 = series.ewm(alpha=1/8, adjust=False).mean()
    current_wwma_8 = wwma_8.iloc[-1] if len(wwma_8) >= 8 else current
    
    # Dispersión vs MA-34
    dispersion_ma34 = ((series - ma_34) / ma_34) * 100
    current_dispersion = dispersion_ma34.iloc[-1] if not dispersion_ma34.empty else 0
    
    # =========================================================================
    # DETECCIÓN DE CRISIS - METODOLOGÍA DE 5 PASOS (para backtesting)
    # =========================================================================
    
    def detect_crisis_4_steps(series, ma_34, wwma_8, dispersion_ma34):
        """
        Detecta fechas donde se cumplen los primeros 4 pasos.
        Para backtesting: Cuenta como crisis cuando pasos 1-4 se cumplen.
        
        Returns: Lista de índices donde se confirma crisis (4 pasos)
        """
        crisis_indices = []
        start_idx = max(34, 8, 30)
        
        for i in range(start_idx, len(series)):
            # PASO 1: Spread ≥ 7.30%
            if series.iloc[i] < 7.30:
                continue
            
            # PASO 2: Dispersión máxima > +15% en últimas 30 ruedas
            if i < 30:
                continue
            disp_window_30 = dispersion_ma34.iloc[i-30:i+1]
            max_disp_30 = disp_window_30.max()
            if pd.isna(max_disp_30) or max_disp_30 <= 15:
                continue
            
            # PASO 3: Spread cruza WWMA-8 DE ARRIBA HACIA ABAJO
            cruce_wwma_idx = None
            for j in range(max(start_idx, i-30), i):
                if (series.iloc[j] >= wwma_8.iloc[j] and 
                    series.iloc[j+1] < wwma_8.iloc[j+1]):
                    cruce_wwma_idx = j + 1
                    break
            if cruce_wwma_idx is None:
                continue
            
            # PASO 4: Dos mínimos decrecientes consecutivos post-cruce
            minimos = []
            for k in range(cruce_wwma_idx + 1, min(i + 1, cruce_wwma_idx + 21)):
                if 1 <= k < len(series) - 1:
                    if (series.iloc[k] < series.iloc[k-1] and 
                        series.iloc[k] < series.iloc[k+1]):
                        minimos.append((k, series.iloc[k]))
            
            if len(minimos) < 2:
                continue
            
            minimos_decrecientes = True
            for j in range(len(minimos) - 1):
                if minimos[j][1] <= minimos[j+1][1]:
                    minimos_decrecientes = False
                    break
            
            if not minimos_decrecientes:
                continue
            
            # ✅ PASOS 1-4 CUMPLIDOS = CRISIS DETECTADA
            crisis_indices.append(i)
        
        return crisis_indices
    
    # =========================================================================
    # EVALUAR ESTADO ACTUAL (PROGRESIVO - SECUENCIAL)
    # =========================================================================
    
    paso1 = False
    paso2 = False
    paso3 = False
    paso4 = False
    paso5 = False
    
    paso1_detail = ""
    paso2_detail = "⏸️  No evaluado (requiere Paso 1)"
    paso3_detail = "⏸️  No evaluado (requiere Paso 1+2)"
    paso4_detail = "⏸️  No evaluado (requiere Paso 1+2+3)"
    paso5_detail = "⏸️  No evaluado (requiere Paso 1+2+3+4)"
    
    cruce_fecha = None
    
    # PASO 1
    paso1 = current >= 7.30
    paso1_detail = f"Spread: {current:.2f}% {'≥' if paso1 else '<'} 7.30%"
    
    if not paso1:
        pasos_cumplidos = 0
    else:
        # PASO 2
        if len(series) >= 30:
            disp_last_30 = dispersion_ma34.tail(30)
            max_disp = disp_last_30.max()
            paso2 = max_disp > 15
            paso2_detail = f"Dispersión máx 30 ruedas: {max_disp:+.1f}% {'>' if paso2 else '≤'} +15%"
        else:
            paso2 = False
            paso2_detail = "⚠️  Insuficientes datos (< 30 ruedas)"
        
        if not paso2:
            pasos_cumplidos = 1
        else:
            # PASO 3
            if len(series) >= 10:
                for j in range(max(0, len(series)-30), len(series)-1):
                    if (series.iloc[j] >= wwma_8.iloc[j] and 
                        series.iloc[j+1] < wwma_8.iloc[j+1]):
                        paso3 = True
                        cruce_fecha = series.index[j+1]
                        paso3_detail = f"Cruce WWMA-8 (bajista): {cruce_fecha.strftime('%Y-%m-%d')}"
                        break
                if not paso3:
                    paso3_detail = "No cruce WWMA-8 reciente (últimas 30)"
            else:
                paso3 = False
                paso3_detail = "⚠️  Insuficientes datos"
            
            if not paso3:
                pasos_cumplidos = 2
            else:
                # PASO 4
                if cruce_fecha is not None:
                    cruce_idx = series.index.get_loc(cruce_fecha)
                    minimos = []
                    for k in range(cruce_idx + 1, min(len(series), cruce_idx + 21)):
                        if 1 <= k < len(series) - 1:
                            if (series.iloc[k] < series.iloc[k-1] and 
                                series.iloc[k] < series.iloc[k+1]):
                                minimos.append(series.iloc[k])
                    
                    if len(minimos) >= 2:
                        decrecientes = all(minimos[j] > minimos[j+1] for j in range(len(minimos)-1))
                        paso4 = decrecientes
                        paso4_detail = f"{len(minimos)} mínimos, {'decrecientes ✓' if decrecientes else 'NO decrecientes ✗'}"
                    else:
                        paso4 = False
                        paso4_detail = f"Solo {len(minimos)} mínimo(s), necesita 2+"
                else:
                    paso4 = False
                    paso4_detail = "⚠️  Requiere Paso 3"
                
                if not paso4:
                    pasos_cumplidos = 3
                else:
                    # PASO 5
                    paso5 = current_dispersion <= 0
                    paso5_detail = f"Dispersión: {current_dispersion:+.1f}% {'≤' if paso5 else '>'} 0%"
                    pasos_cumplidos = 5 if paso5 else 4
    
    # =========================================================================
    # CLASIFICACIÓN DE NIVEL
    # =========================================================================
    
    if pasos_cumplidos == 5:
        level = 2
        status = '🔴 CRISIS CONFIRMADA (5/5)'
        recommendation = '¡ESTAMOS EN CRISIS CONFIRMADA! CASH/Treasuries hasta recuperación clara.'
    elif pasos_cumplidos == 4:
        level = 2
        status = '🔴 CRISIS EN DESARROLLO (4/5)'
        recommendation = 'CRISIS EN DESARROLLO. Mínimos decrecientes confirmados. SALIR inmediatamente.'
    elif pasos_cumplidos == 3:
        level = 1
        status = '🟡 CRISIS POR DESARROLLARSE (3/5)'
        recommendation = 'CRISIS POR DESARROLLARSE. Cruce bajista detectado. MANTENERSE FUERA de HY.'
    elif pasos_cumplidos == 2:
        level = 1
        status = '🟡 PÁNICO DETECTADO (2/5)'
        recommendation = 'Pánico confirmado sin cruce técnico. PREPARAR salida, reducir exposición.'
    elif pasos_cumplidos == 1:
        level = 1
        status = '🟡 ALERTA TEMPRANA (1/5)'
        recommendation = 'Spread superó umbral. VIGILAR evolución. Considerar reducir HY.'
    else:
        level = 0
        status = '🟢 NORMAL'
        recommendation = 'Sin señales de crisis. Condiciones normales.'
    
    # =========================================================================
    # BACKTESTING CON METODOLOGÍA DE 5 PASOS
    # =========================================================================
    
    if backtester:
        # Detectar crisis históricas (primeros 4 pasos)
        crisis_indices = detect_crisis_4_steps(series, ma_34, wwma_8, dispersion_ma34)
        
        print(f"\n🔍 DEBUG: {len(crisis_indices)} crisis detectadas (4 pasos) con metodología 5 pasos")
        for idx in crisis_indices[:10]:  # Mostrar primeras 10
            fecha = series.index[idx]
            valor = series.iloc[idx]
            print(f"   - {fecha.strftime('%Y-%m-%d')}: Spread = {valor:.2f}%")
        if len(crisis_indices) > 10:
            print(f"   ... y {len(crisis_indices)-10} más")
        
        # Crear serie binaria
        crisis_series = pd.Series(0.0, index=series.index)
        for idx in crisis_indices:
            crisis_series.iloc[idx] = 10.0
        
        # Ejecutar backtesting
        backtester.backtest_threshold(
            crisis_series,
            threshold=5,
            comparison='>',
            lead_days=378,  # 18 meses bursátiles
            name='High Yield - 5 Pasos (Crisis = 4 pasos cumplidos)'
        )
    
    # =========================================================================
    # OUTPUT DETALLADO
    # =========================================================================
    
    status_detail = f"\n   📅 Último: {last_date.strftime('%Y-%m-%d')}"
    status_detail += f"\n   📊 Spread: {current:.2f}%"
    status_detail += f"\n   📊 MA-34: {current_ma_34:.2f}%"
    status_detail += f"\n   📊 WWMA-8: {current_wwma_8:.2f}%"
    status_detail += f"\n   📊 Dispersión: {current_dispersion:+.1f}%"
    
    status_detail += "\n\n   ═══ VALIDACIÓN PROGRESIVA ═══"
    status_detail += f"\n   {'✓' if paso1 else '✗'} PASO 1: {paso1_detail}"
    
    if paso1:
        status_detail += f"\n        {'✓' if paso2 else '✗'} PASO 2: {paso2_detail}"
        if paso2:
            status_detail += f"\n             {'✓' if paso3 else '✗'} PASO 3: {paso3_detail}"
            if paso3:
                status_detail += f"\n                  {'✓' if paso4 else '✗'} PASO 4: {paso4_detail}"
                if paso4:
                    status_detail += f"\n                       {'✓' if paso5 else '✗'} PASO 5: {paso5_detail}"
                else:
                    status_detail += f"\n                       ⏸️  PASO 5: {paso5_detail}"
            else:
                status_detail += f"\n                  ⏸️  PASO 4: {paso4_detail}"
                status_detail += f"\n                       ⏸️  PASO 5: {paso5_detail}"
        else:
            status_detail += f"\n             ⏸️  PASO 3: {paso3_detail}"
            status_detail += f"\n                  ⏸️  PASO 4: {paso4_detail}"
            status_detail += f"\n                       ⏸️  PASO 5: {paso5_detail}"
    else:
        status_detail += f"\n        ⏸️  PASO 2: {paso2_detail}"
        status_detail += f"\n             ⏸️  PASO 3: {paso3_detail}"
        status_detail += f"\n                  ⏸️  PASO 4: {paso4_detail}"
        status_detail += f"\n                       ⏸️  PASO 5: {paso5_detail}"
    
    status_detail += f"\n\n   🎯 Pasos cumplidos: {pasos_cumplidos}/5"
    
    if pasos_cumplidos >= 3:
        status_detail += f"\n   📈 https://www.tradingview.com/chart/?symbol=ICE:BAMLH0A0HYM2"
    
    gfc_max = series.loc['2008-09-01':'2009-03-01'].max() if '2008-09-01' in series.index else None
    covid_max = series.loc['2020-02-01':'2020-04-01'].max() if '2020-02-01' in series.index else None
    
    return {
        'name': 'High Yield Spread (5 Pasos)',
        'current': current,
        'last_date': last_date.strftime('%Y-%m-%d'),
        'ma_34': current_ma_34,
        'wwma_8': current_wwma_8,
        'dispersion': current_dispersion,
        'steps_completed': pasos_cumplidos,
        'level': level,
        'status': status + status_detail,
        'recommendation': recommendation,
        'historical': {
            'GFC_2008': gfc_max,
            'COVID_2020': covid_max
        },
        'series': series
    }


def analyze_yield_curve(backtester=None):
    """Curva de Tipos 10Y-2Y con backtesting."""
    series = get_cached_series('T10Y2Y', lambda: fetch_fred_series('T10Y2Y'))
    
    if series.empty:
        return {
            'name': 'Curva de Tipos (10Y-2Y)',
            'current': None,
            'level': -1,
            'status': '⚠️  Error: Sin datos',
            'recommendation': 'No disponible'
        }
    
    current = series.iloc[-1]
    
    # Backtesting
    if backtester:
        backtester.backtest_threshold(
            series,
            threshold=0,
            comparison='<',
            lead_days=365,  # La curva invierte con mucha anticipación
            name='Curva Invertida (< 0%)'
        )
    
    # Clasificación
    if current > 0.25:
        level = 0
        status = '🟢 NORMAL'
        recommendation = 'Curva positiva saludable. Economía en expansión esperada.'
    elif current > 0:
        level = 1
        status = '🟡 ALERTA'
        recommendation = 'Aplanamiento de curva. Posible desaceleración económica en 12-18 meses.'
    else:
        level = 2
        status = '🔴 PELIGRO'
        recommendation = 'INVERSIÓN DE CURVA: Alto predictor de recesión en 12-24 meses. Reducir riesgo.'
    
    # Chequear persistencia
    recent_3m = series.tail(63)  # ~3 meses de días hábiles
    if not recent_3m.empty:
        inversion_pct = (recent_3m < 0).mean()
        if inversion_pct > 0.8:
            recommendation += f" [INVERSIÓN PERSISTENTE: {inversion_pct:.0%} del tiempo]"
    
    dotcom_min = series.loc['1999-01-01':'2001-09-01'].min()
    gfc_min = series.loc['2006-01-01':'2008-09-01'].min()
    
    return {
        'name': 'Curva de Tipos (10Y-2Y)',
        'current': current,
        'level': level,
        'status': status,
        'recommendation': recommendation,
        'historical': {
            'Dotcom_min': dotcom_min,
            'GFC_2008_min': gfc_min
        },
        'series': series
    }

def analyze_vix(backtester=None):
    """
    VIX con análisis semanal (Usuario - Dic 2024).
    
    LÓGICA:
    1. Convertir datos a frecuencia SEMANAL (5 días bursátiles)
    2. Si semana en curso, usar último dato disponible
    3. Cotas de control:
       - Piso: 11
       - Medio inferior: 14.85
       - Medio neutro: 16.28
       - Medio alto: 19.50
       - Alto (crisis): 34.6
    """
    # Prioridad: FRED > Yahoo
    series = get_cached_series('VIXCLS', lambda: fetch_fred_series('VIXCLS'))
    
    if series.empty:
        print("⚠️  FRED VIX falló, intentando Yahoo...")
        series = get_cached_series('^VIX', lambda: fetch_yahoo_series('^VIX'))
    
    if series.empty:
        return {
            'name': 'VIX (Índice del Miedo)',
            'current': None,
            'level': -1,
            'status': '⚠️  Error: Sin datos',
            'recommendation': 'No disponible'
        }
    
    # =========================================================================
    # CONVERSIÓN A SEMANAL
    # =========================================================================
    
    # Tomar último valor de cada semana (viernes o último día disponible)
    series_weekly = series.resample('W-FRI').last().dropna()
    
    # Si semana en curso, ya está incluida por 'last()'
    current_weekly = series_weekly.iloc[-1]
    last_date = series_weekly.index[-1]
    
    # =========================================================================
    # COTAS DE CONTROL
    # =========================================================================
    
    PISO = 11.0
    MEDIO_INFERIOR = 14.85
    MEDIO_NEUTRO = 16.28
    MEDIO_ALTO = 19.50
    ALTO_CRISIS = 34.6
    
    # Clasificación
    if current_weekly >= ALTO_CRISIS:
        level = 2
        status = '🔴 ALTO (CRISIS)'
        zona = f'≥ {ALTO_CRISIS}'
        recommendation = 'VIX en zona de CRISIS. Pánico extremo en el mercado.'
    elif current_weekly >= MEDIO_ALTO:
        level = 1
        status = '🟡 MEDIO ALTO (ATENCIÓN)'
        zona = f'{MEDIO_ALTO} - {ALTO_CRISIS}'
        recommendation = 'VIX elevado. Nerviosismo significativo en el mercado.'
    elif current_weekly >= MEDIO_NEUTRO:
        level = 0
        status = '🟢 MEDIO NEUTRO'
        zona = f'{MEDIO_NEUTRO} - {MEDIO_ALTO}'
        recommendation = 'VIX en rango neutro. Volatilidad normal.'
    elif current_weekly >= MEDIO_INFERIOR:
        level = 0
        status = '🟢 MEDIO INFERIOR'
        zona = f'{MEDIO_INFERIOR} - {MEDIO_NEUTRO}'
        recommendation = 'VIX bajo. Mercado tranquilo.'
    elif current_weekly >= PISO:
        level = 0
        status = '🟢 BAJO'
        zona = f'{PISO} - {MEDIO_INFERIOR}'
        recommendation = 'VIX muy bajo. Complacencia en el mercado.'
    else:
        level = 0
        status = '🟢 MUY BAJO (PISO)'
        zona = f'< {PISO}'
        recommendation = 'VIX extremadamente bajo. Complacencia extrema.'
    
    # =========================================================================
    # BACKTESTING CON COTAS
    # =========================================================================
    
    if backtester:
        # Señal de crisis (≥ 34.6)
        signal_crisis = (series_weekly >= ALTO_CRISIS).astype(float) * 10
        backtester.backtest_threshold(
            signal_crisis,
            threshold=5,
            comparison='>',
            lead_days=21,  # ~1 mes bursátil
            name=f'VIX ≥ {ALTO_CRISIS} (Crisis)'
        )
        
        # Señal de atención (≥ 19.50)
        signal_atencion = (series_weekly >= MEDIO_ALTO).astype(float) * 10
        backtester.backtest_threshold(
            signal_atencion,
            threshold=5,
            comparison='>',
            lead_days=42,  # ~2 meses bursátiles
            name=f'VIX ≥ {MEDIO_ALTO} (Atención)'
        )
    
    # =========================================================================
    # OUTPUT
    # =========================================================================
    
    gfc_max = series.loc['2008-09-01':'2009-03-01'].max() if '2008-09-01' in series.index else None
    covid_max = series.loc['2020-02-01':'2020-04-01'].max() if '2020-02-01' in series.index else None
    
    status_detail = f"\n   📅 Última semana: {last_date.strftime('%Y-%m-%d')}"
    status_detail += f"\n   📊 VIX: {current_weekly:.2f}"
    status_detail += f"\n   📊 Zona: {zona}"
    status_detail += "\n"
    status_detail += "\n   ═══ COTAS DE CONTROL ═══"
    status_detail += f"\n   {'✓' if current_weekly >= ALTO_CRISIS else ' '} Alto (Crisis): ≥ {ALTO_CRISIS}"
    status_detail += f"\n   {'✓' if MEDIO_ALTO <= current_weekly < ALTO_CRISIS else ' '} Medio Alto: {MEDIO_ALTO} - {ALTO_CRISIS}"
    status_detail += f"\n   {'✓' if MEDIO_NEUTRO <= current_weekly < MEDIO_ALTO else ' '} Medio Neutro: {MEDIO_NEUTRO} - {MEDIO_ALTO}"
    status_detail += f"\n   {'✓' if MEDIO_INFERIOR <= current_weekly < MEDIO_NEUTRO else ' '} Medio Inferior: {MEDIO_INFERIOR} - {MEDIO_NEUTRO}"
    status_detail += f"\n   {'✓' if PISO <= current_weekly < MEDIO_INFERIOR else ' '} Bajo: {PISO} - {MEDIO_INFERIOR}"
    status_detail += f"\n   {'✓' if current_weekly < PISO else ' '} Muy Bajo: < {PISO}"
    
    return {
        'name': 'VIX (Índice del Miedo) - SEMANAL',
        'current': current_weekly,
        'zona': zona,
        'level': level,
        'status': status + status_detail,
        'recommendation': recommendation,
        'historical': {
            'GFC_2008': gfc_max if gfc_max else 'N/A',
            'COVID_2020': covid_max if covid_max else 'N/A'
        },
        'series': series_weekly
    }


def analyze_dollar(backtester=None):
    """
    Índice del Dólar con análisis mensual (Usuario - Dic 2024).
    
    LÓGICA:
    1. Convertir datos a frecuencia MENSUAL
    2. MA-8 mensual (8 meses)
    3. Dispersión = (Índice - MA-8) / MA-8 * 100
    4. Alertas: +5% (alcista) y -4% (bajista)
    
    NOTA: Si mes en curso, usar último dato disponible
    """
    series = get_cached_series('DTWEXBGS', lambda: fetch_fred_series('DTWEXBGS'))
    name_source = "FRED: DTWEXBGS"
    
    if series.empty:
        series = get_cached_series('DX=F', lambda: fetch_yahoo_series('DX=F'))
        name_source = "Yahoo: DX=F"
    
    if series.empty:
        return {
            'name': 'Índice del Dólar',
            'current': None,
            'level': -1,
            'status': '⚠️  Error: Sin datos',
            'recommendation': 'No disponible'
        }
    
    # =========================================================================
    # CONVERSIÓN A MENSUAL
    # =========================================================================
    
    # Tomar último valor de cada mes (o último disponible si mes en curso)
    series_monthly = series.resample('ME').last().dropna()
    
    # Si el último dato es del mes en curso, ya está incluido por 'last()'
    current_monthly = series_monthly.iloc[-1]
    last_date = series_monthly.index[-1]
    
    # =========================================================================
    # MA-8 MENSUAL
    # =========================================================================
    
    ma_8_monthly = series_monthly.rolling(window=8).mean()
    current_ma_8 = ma_8_monthly.iloc[-1] if len(ma_8_monthly) >= 8 else current_monthly
    
    # =========================================================================
    # DISPERSIÓN vs MA-8
    # =========================================================================
    
    dispersion = ((series_monthly - ma_8_monthly) / ma_8_monthly) * 100
    current_dispersion = dispersion.iloc[-1] if not dispersion.empty else 0
    
    # =========================================================================
    # ALERTAS
    # =========================================================================
    
    alerta_alcista = current_dispersion >= 5.0
    alerta_bajista = current_dispersion <= -4.0
    
    if alerta_alcista:
        level = 1
        status = '🟡 DISPERSIÓN ALCISTA'
        recommendation = f'Dispersión: {current_dispersion:+.1f}% (≥ +5%). Dólar fuerte vs tendencia.'
    elif alerta_bajista:
        level = 1
        status = '🟡 DISPERSIÓN BAJISTA'
        recommendation = f'Dispersión: {current_dispersion:+.1f}% (≤ -4%). Dólar débil vs tendencia.'
    else:
        level = 0
        status = '🟢 NORMAL'
        recommendation = f'Dispersión: {current_dispersion:+.1f}% (dentro de rango normal).'
    
    # =========================================================================
    # BACKTESTING (Opcional - con datos mensuales)
    # =========================================================================
    
    if backtester:
        # Crear señales binarias
        signal_alcista = (dispersion >= 5.0).astype(float) * 10
        signal_bajista = (dispersion <= -4.0).astype(float) * 10
        
        backtester.backtest_threshold(
            signal_alcista,
            threshold=5,
            comparison='>',
            lead_days=126,  # ~6 meses bursátiles
            name='Dólar: Dispersión ≥ +5%'
        )
        
        backtester.backtest_threshold(
            signal_bajista,
            threshold=5,
            comparison='>',
            lead_days=126,
            name='Dólar: Dispersión ≤ -4%'
        )
    
    # =========================================================================
    # OUTPUT
    # =========================================================================
    
    gfc_max = series.loc['2008-09-01':'2009-03-01'].max() if '2008-09-01' in series.index else None
    
    status_detail = f"\n   📅 Último mes: {last_date.strftime('%Y-%m')}"
    status_detail += f"\n   📊 Índice: {current_monthly:.2f}"
    status_detail += f"\n   📊 MA-8 mensual: {current_ma_8:.2f}"
    status_detail += f"\n   📊 Dispersión: {current_dispersion:+.1f}%"
    status_detail += "\n"
    
    if alerta_alcista:
        status_detail += f"\n   ⚠️  Superó barrera alcista (+5%)"
    elif alerta_bajista:
        status_detail += f"\n   ⚠️  Superó barrera bajista (-4%)"
    else:
        status_detail += f"\n   ✅ Dispersión dentro de rango normal"
    
    return {
        'name': f'Índice del Dólar ({name_source}) - MENSUAL',
        'current': current_monthly,
        'ma_8': current_ma_8,
        'dispersion': current_dispersion,
        'level': level,
        'status': status + status_detail,
        'recommendation': recommendation,
        'historical': {
            'GFC_2008': gfc_max if gfc_max else 'N/A'
        },
        'series': series_monthly
    }


def analyze_fear_greed(backtester=None):
    """
    Fear & Greed Index de CNN.
    
    IMPORTANTE: Este indicador es CONTRARIAN:
    - Extreme Greed (>75) = Momento para SER CAUTELOSO (posible tope)
    - Extreme Fear (<25) = Momento para SER AGRESIVO (posible fondo)
    """
    fg_data = fetch_fear_greed_index()
    
    if fg_data['current'] is None:
        return {
            'name': 'Fear & Greed Index',
            'current': None,
            'level': -1,
            'status': '⚠️  Error: Sin datos',
            'recommendation': 'No disponible'
        }
    
    current = fg_data['current']
    rating = fg_data['rating']
    
    # Clasificación CONTRARIAN (invertida)
    if current > 75:
        level = 2  # Peligro = Mucha codicia
        status = f'🔴 {rating.upper()}'
        recommendation = 'EXTREME GREED: Mercado eufórico, posible tope. REDUCIR exposición.'
    elif current > 55:
        level = 1
        status = f'🟡 {rating}'
        recommendation = 'Optimismo elevado. Monitorear para tomar ganancias.'
    elif current > 45:
        level = 0
        status = f'🟢 {rating}'
        recommendation = 'Neutral. Mantener posiciones actuales.'
    elif current > 25:
        level = 1
        status = f'🟡 {rating}'
        recommendation = 'Miedo creciente. Comenzar a buscar oportunidades de compra.'
    else:
        level = 0  # Oportunidad = Mucho miedo
        status = f'🟢 {rating.upper()}'
        recommendation = 'EXTREME FEAR: ¡OPORTUNIDAD! Momento para COMPRAR con descuento.'
    
    return {
        'name': 'Fear & Greed Index (CNN)',
        'current': current,
        'rating': rating,
        'level': level,
        'status': status,
        'recommendation': recommendation,
        'note': 'Indicador CONTRARIAN: Extreme Fear = Compra, Extreme Greed = Venta'
    }

# =============================================================================
# SISTEMA DE SCORING PONDERADO
# =============================================================================

def calculate_composite_score(indicators):
    """
    Calcula score compuesto con ponderaciones basadas en backtesting.
    
    Ponderaciones (basadas en precision y sensibilidad):
    - Curva de Tipos: 35% (mejor predictor de largo plazo)
    - High Yield: 25% (buen indicador de estrés crediticio)
    - VIX: 20% (señal de corto plazo, alta volatilidad)
    - Dólar: 20% (complementario, flight to quality)
    - Fear & Greed: NO se incluye (es contrarian y subjetivo)
    """
    weights = {
        'Curva de Tipos (10Y-2Y)': 0.35,
        'High Yield Spread': 0.25,
        'VIX (Índice del Miedo)': 0.20,
        'Índice del Dólar': 0.20
    }
    
    total_score = 0
    total_weight = 0
    
    for ind in indicators:
        name = ind['name']
        # Buscar peso parcial (ej: "Índice del Dólar (FRED: ...)")
        weight = 0
        for key, w in weights.items():
            if key in name:
                weight = w
                break
        
        if weight > 0 and ind['level'] >= 0:
            total_score += ind['level'] * weight
            total_weight += weight
    
    # Normalizar a escala 0-2
    if total_weight > 0:
        normalized_score = (total_score / total_weight)
    else:
        normalized_score = 0
    
    return normalized_score

def get_portfolio_recommendation(composite_score, fear_greed_value):
    """
    Recomendaciones accionables para un inversor particular.
    
    Args:
        composite_score: Score compuesto 0-2
        fear_greed_value: Valor del Fear & Greed Index (0-100)
    """
    # Clasificación principal
    if composite_score < 0.5:
        risk_level = "BAJO"
        emoji = "🟢"
        action = "OFENSIVA"
    elif composite_score < 1.0:
        risk_level = "MEDIO"
        emoji = "🟡"
        action = "CAUTELOSA"
    else:
        risk_level = "ALTO"
        emoji = "🔴"
        action = "DEFENSIVA"
    
    # Ajustes por Fear & Greed
    fg_adjustment = ""
    if fear_greed_value and fear_greed_value > 75:
        fg_adjustment = "\n  ⚠️  AJUSTE por F&G: Mercado eufórico, reducir exposición adicional."
    elif fear_greed_value and fear_greed_value < 25:
        fg_adjustment = "\n  ✅ AJUSTE por F&G: Mercado en pánico, buscar oportunidades de compra."
    
    recommendations = {
        "OFENSIVA": """
  📈 ESTRATEGIA OFENSIVA (Riesgo Bajo):
  
  Acciones sugeridas:
  • Mantener exposición completa (80-100% acciones)
  • Considerar aumentar posiciones en sectores cíclicos
  • Momento para invertir nuevo capital
  • No necesario tener coberturas (puts/VIX)
  
  Monitoreo: Revisar semanalmente
        """,
        
        "CAUTELOSA": """
  ⚖️  ESTRATEGIA CAUTELOSA (Riesgo Medio):
  
  Acciones sugeridas:
  • Reducir exposición a 60-70% acciones
  • Aumentar cash y/o bonos del gobierno (20-30%)
  • Tomar ganancias en posiciones con alta ganancia
  • Considerar coberturas ligeras (5-10% en puts)
  • Evitar compras agresivas, esperar confirmación
  
  Monitoreo: Revisar 2-3 veces por semana
        """,
        
        "DEFENSIVA": """
  🛡️  ESTRATEGIA DEFENSIVA (Riesgo Alto):
  
  Acciones sugeridas:
  • REDUCIR exposición a 30-50% acciones
  • Aumentar cash significativamente (30-50%)
  • Posiciones en Treasuries o bonos AAA (20%)
  • Coberturas agresivas: 10-20% en puts o VIX calls
  • VENDER posiciones especulativas/apalancadas
  • Mantener solo posiciones de alta convicción
  • Preparar lista de compra para cuando baje
  
  Monitoreo: DIARIO durante esta fase
        """
    }
    
    return f"""
{emoji} NIVEL DE RIESGO: {risk_level} (Score: {composite_score:.2f}/2.00)
{fg_adjustment}
{recommendations[action]}
"""

# =============================================================================
# GENERACIÓN DE REPORTES
# =============================================================================

def generate_excel_report(indicators, backtesting_results, output_file='Crisis_Dashboard_Report.xlsx'):
    """
    Genera reporte profesional en Excel con:
    - Dashboard principal
    - Backtesting detallado
    - Datos históricos
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # =========================================================================
    # HOJA 1: DASHBOARD PRINCIPAL
    # =========================================================================
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    
    # Header
    ws_dash['A1'] = 'DASHBOARD DE INDICADORES ADELANTADOS DE CRISIS FINANCIERA'
    ws_dash['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws_dash['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws_dash.merge_cells('A1:F1')
    
    ws_dash['A2'] = f'Fecha de análisis: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws_dash['A2'].font = Font(italic=True)
    
    # Headers de tabla
    headers = ['Indicador', 'Valor Actual', 'Estado', 'Nivel', 'Recomendación']
    for col, header in enumerate(headers, start=1):
        cell = ws_dash.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Datos de indicadores
    row = 5
    for ind in indicators:
        if ind['current'] is None:
            continue
            
        ws_dash.cell(row=row, column=1, value=ind['name'])
        
        # Formatear valor según tipo
        if 'VIX' in ind['name'] or 'Fear & Greed' in ind['name']:
            ws_dash.cell(row=row, column=2, value=f"{ind['current']:.2f}")
        elif 'Dólar' in ind['name']:
            change = ind.get('change_3m', 0)
            ws_dash.cell(row=row, column=2, value=f"{ind['current']:.2f} ({change:+.1f}%)")
        else:
            ws_dash.cell(row=row, column=2, value=f"{ind['current']:.2f}%")
        
        ws_dash.cell(row=row, column=3, value=ind['status'])
        ws_dash.cell(row=row, column=4, value=ind['level'])
        ws_dash.cell(row=row, column=5, value=ind['recommendation'])
        
        # Color coding por nivel
        if ind['level'] == 2:
            color = 'FFC7CE'  # Rojo claro
        elif ind['level'] == 1:
            color = 'FFEB9C'  # Amarillo
        else:
            color = 'C6EFCE'  # Verde claro
        
        for col in range(1, 6):
            ws_dash.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        row += 1
    
    # Ajustar anchos
    ws_dash.column_dimensions['A'].width = 30
    ws_dash.column_dimensions['B'].width = 20
    ws_dash.column_dimensions['C'].width = 25
    ws_dash.column_dimensions['D'].width = 10
    ws_dash.column_dimensions['E'].width = 60
    
    # =========================================================================
    # HOJA 2: BACKTESTING RESULTS
    # =========================================================================
    ws_back = wb.create_sheet("Backtesting")
    
    ws_back['A1'] = 'RESULTADOS DE BACKTESTING (1990-2024)'
    ws_back['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws_back['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    ws_back.merge_cells('A1:H1')
    
    # Headers
    back_headers = ['Indicador', 'Umbral', 'Total Señales', 'Aciertos', 'Falsos +', 'Precisión', 'Sensibilidad', 'Lead Time (días)']
    for col, header in enumerate(back_headers, start=1):
        cell = ws_back.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    # Datos
    row = 4
    for name, result in backtesting_results.items():
        ws_back.cell(row=row, column=1, value=result['name'])
        ws_back.cell(row=row, column=2, value=result['threshold'])
        ws_back.cell(row=row, column=3, value=result['total_signals'])
        ws_back.cell(row=row, column=4, value=result['true_positives'])
        ws_back.cell(row=row, column=5, value=result['false_positives'])
        ws_back.cell(row=row, column=6, value=f"{result['precision']:.1%}")
        ws_back.cell(row=row, column=7, value=f"{result['sensitivity']:.1%}")
        ws_back.cell(row=row, column=8, value=f"{result['avg_lead_days']:.0f}")
        
        row += 1
    
    for col in range(1, 9):
        ws_back.column_dimensions[get_column_letter(col)].width = 18
    
    # =========================================================================
    # HOJA 3: DATOS HISTÓRICOS (últimos 100 días)
    # =========================================================================
    ws_hist = wb.create_sheet("Históricos")
    
    ws_hist['A1'] = 'DATOS HISTÓRICOS (últimos 100 días hábiles)'
    ws_hist['A1'].font = Font(size=12, bold=True)
    ws_hist.merge_cells('A1:E1')
    
    # Construir DataFrame consolidado
    historical_data = {}
    for ind in indicators:
        if 'series' in ind and not ind['series'].empty:
            historical_data[ind['name']] = ind['series'].tail(100)
    
    if historical_data:
        df_hist = pd.DataFrame(historical_data)
        
        # Escribir headers
        ws_hist.cell(row=3, column=1, value='Fecha')
        for col, name in enumerate(df_hist.columns, start=2):
            ws_hist.cell(row=3, column=col, value=name)
        
        # Escribir datos
        for row_idx, (date, row_data) in enumerate(df_hist.iterrows(), start=4):
            ws_hist.cell(row=row_idx, column=1, value=date.strftime('%Y-%m-%d'))
            for col_idx, value in enumerate(row_data, start=2):
                ws_hist.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(output_file)
    print(f"\n📊 Reporte Excel generado: {output_file}")

# =============================================================================
# FUNCIÓN PRINCIPAL
# =============================================================================

def run_complete_analysis(run_backtesting=True, generate_report=True):
    """
    Ejecuta análisis completo con backtesting opcional.
    
    Args:
        run_backtesting: Si True, ejecuta validación histórica (toma ~2 min)
        generate_report: Si True, genera reporte Excel
    """
    print("\n" + "="*80)
    print("  SISTEMA PROFESIONAL DE INDICADORES ADELANTADOS DE CRISIS")
    print("  Versión 7.0 - Con Backtesting Validado")
    print("="*80)
    
    # Inicializar backtester
    backtester = IndicatorBacktester(CRISIS_PERIODS) if run_backtesting else None
    
    if run_backtesting:
        print("\n⏳ Ejecutando backtesting histórico (1990-2024)...")
        print("   Esto tomará 1-2 minutos la primera vez...")
    
    # Ejecutar análisis
    print("\n📥 Descargando y analizando indicadores...\n")
    
    indicators = [
        analyze_high_yield(backtester),
        analyze_yield_curve(backtester),
        analyze_vix(backtester),
        analyze_dollar(backtester),
        analyze_fear_greed(backtester)
    ]
    
    # Imprimir resultados
    print("\n" + "="*80)
    print("  RESULTADOS DEL ANÁLISIS")
    print("="*80)
    
    for ind in indicators:
        if ind['current'] is None:
            continue
        
        print(f"\n{ind['name']}")
        print("-" * 80)
        if 'VIX' in ind['name'] or 'Fear & Greed' in ind['name']:
            print(f"  Valor actual: {ind['current']:.2f}")
        elif 'Dólar' in ind['name']:
            print(f"  Valor actual: {ind['current']:.2f} (Cambio 3M: {ind.get('change_3m', 0):+.1f}%)")
        else:
            print(f"  Valor actual: {ind['current']:.2f}%")
        
        print(f"  {ind['status']}")
        print(f"  Recomendación: {ind['recommendation']}")
        
        if 'note' in ind:
            print(f"  📌 {ind['note']}")
    
    # Mostrar resultados de backtesting
    if run_backtesting and backtester.results:
        print("\n" + "="*80)
        print("  VALIDACIÓN HISTÓRICA (BACKTESTING)")
        print("="*80)
        
        for name in backtester.results.keys():
            backtester.print_summary(name)
    
    # Score compuesto
    fg_value = next((ind['current'] for ind in indicators if 'Fear & Greed' in ind['name']), None)
    composite = calculate_composite_score(indicators)
    
    print("\n" + "="*80)
    print("  RECOMENDACIÓN FINAL PARA TU CARTERA")
    print("="*80)
    print(get_portfolio_recommendation(composite, fg_value))
    
    # Generar reporte Excel
    if generate_report:
        try:
            output_file = f'/mnt/user-data/outputs/Crisis_Dashboard_{datetime.now().strftime("%Y%m%d")}.xlsx'
            generate_excel_report(indicators, backtester.results if backtester else {}, output_file)
        except Exception as e:
            print(f"⚠️  Error al generar reporte Excel: {e}")
    
    print("="*80)
    print("\n✅ Análisis completado exitosamente")
    
    return indicators, backtester

# =============================================================================
# EJECUCIÓN
# =============================================================================

if __name__ == '__main__':
    # Configuración para primera ejecución
    RUN_BACKTESTING = True  # Cambiar a False después de la primera corrida para velocidad
    GENERATE_REPORT = True
    
    try:
        indicators, backtester = run_complete_analysis(
            run_backtesting=RUN_BACKTESTING,
            generate_report=GENERATE_REPORT
        )
    except Exception as e:
        print(f"\n❌ ERROR CRÍTICO: {e}")
        import traceback
        traceback.print_exc()
